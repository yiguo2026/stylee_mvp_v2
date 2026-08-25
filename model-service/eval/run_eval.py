#!/usr/bin/env python3
"""Stylee 离线搭配能力评测主入口。

用法:
    python run_eval.py --catalog catalog.sample.json --queries queries.json \
        --provider mock --out results/

流程: 读取单品池(catalog) + query 集 -> 逐条调用 model-service 推荐链路(默认 mock provider)
      -> 产出 results/outfits.jsonl / review_sheet.csv / review_sheet.xlsx /
         review_sheet_personalization.csv

不修改 model-service 源码；通过 import 复用其 adapter + pipeline + provider。
切换真实模型只需 --provider deepseek|qwen（并在环境变量里配置对应 key），代码零改动。
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys

import eval_lib as E

WEIGHTS = {"intent": 0.30, "aesthetic": 0.25, "personalization": 0.20,
           "wearability": 0.15, "creativity": 0.10}

WEIGHTED_HEADER = "加权总分(=0.3意图+0.25审美+0.2个性化+0.15实穿+0.1创意)"

REVIEW_HEADER = [
    "query_id", "query文本", "场景", "风格", "色系", "季节/温度",
    "上装", "下装", "外套", "鞋", "包", "配饰",
    "意图匹配(1-5)", "审美质感(1-5)", "个性化深度(1-5)", "实穿性(1-5)", "创意惊喜(1-5)",
    WEIGHTED_HEADER, "评审备注",
]

GUIDE_LINES = [
    "【评审说明】本表左侧为模型生成的整套搭配，右侧为 5 个维度打分（每项 1-5 分，整数）。",
    "1) 意图匹配(权重30%)：是否满足 query 核心诉求(场景/风格/指定单品)。5=完美命中；3=部分命中主要诉求；1=完全跑题。",
    "2) 审美质感(权重25%)：整体美感/和谐度/高级感。5=接近时尚基准；3=中等；1=明显违和。",
    "3) 个性化深度(权重20%)：是否利用用户画像做差异化。5=同query不同画像有显著差异；3=部分个性化；1=千人一面。(需成对双画像对比，详见 personalization 子表)",
    "4) 实穿性(权重15%)：真的会穿出门吗(天气/舒适/场景)。5=完全能穿；3=有一定阻力；1=概念稿。",
    "5) 创意/惊喜感(权重10%)：有无巧妙且不离谱的组合。5=有巧思且不离谱；3=常规组合；1=毫无新意或过于离谱。",
    "加权总分 = 0.30*意图 + 0.25*审美 + 0.20*个性化 + 0.15*实穿 + 0.10*创意（xlsx 自动计算；填完 1-5 即出分）。",
]


def build_rows(queries, catalog, provider, n):
    """跑全部 query，返回 (jsonl_records, review_rows, per_query_top)。"""
    cat_idx = E.catalog_index(catalog)
    jsonl_records = []
    review_rows = []
    per_query_top = {}   # query_id -> extracted top outfit（供个性化子表）

    for q in queries:
        payload, ctx, result = E.run_one(q, catalog, provider, n=n)
        extracted = [E.extract_outfit(o, cat_idx) for o in result.outfits]
        per_query_top[q["query_id"]] = {
            "query": q,
            "top": extracted[0] if extracted else None,
        }

        # jsonl：一行一 query，含模型选择、理由、原始响应
        jsonl_records.append({
            "query_id": q["query_id"],
            "text": q["text"],
            "labels": {k: q.get(k, "") for k in
                       ("scenario", "style", "color_system", "season",
                        "temp_range", "gender", "special", "profile_variant")},
            "payload_weather": payload["weather"],
            "payload_profile": payload["profile"],
            "provider": provider.name,
            "model_version": result.model_version,
            "num_outfits": len(result.outfits),
            "outfits": [
                {
                    "rank": i + 1,
                    "slot_selection": {s["slot"]: s.get("item_id") for s in eo["slots"]},
                    "slots": eo["slots"],
                    "display_columns": eo["display_columns"],
                    "reasoning": eo["reasoning"],
                    "confidence": eo["confidence"],
                    "scores": eo["scores"],
                }
                for i, eo in enumerate(extracted)
            ],
            "trace": result.trace,
        })

        # review：默认对 top-1（主推荐方案）做人评
        top = extracted[0] if extracted else None
        cols = top["display_columns"] if top else {c: "—" for c in E.SLOT_COLUMNS}
        review_rows.append([
            q["query_id"], q["text"], q.get("scenario", ""), q.get("style", ""),
            q.get("color_system", ""),
            f"{q.get('season','')} / {q.get('temp_range','')}",
            cols["上装"], cols["下装"], cols["外套"], cols["鞋"], cols["包"], cols["配饰"],
            "", "", "", "", "",   # 5 个待打分维度
            "",                     # 加权总分（xlsx 用公式；csv 留空）
            "",                     # 评审备注
        ])

    return jsonl_records, review_rows, per_query_top


def write_jsonl(path, records):
    with open(path, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def write_review_csv(path, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        for line in GUIDE_LINES:
            w.writerow([line])
        w.writerow([])
        w.writerow(REVIEW_HEADER)
        for r in rows:
            w.writerow(r)


def write_review_xlsx(path, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception as e:  # pragma: no cover
        print(f"[warn] 无 openpyxl，跳过 xlsx：{e}")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "推荐评审"

    # 顶部评审说明
    r = 1
    for line in GUIDE_LINES:
        ws.cell(row=r, column=1, value=line)
        ws.cell(row=r, column=1).font = Font(size=10, color="444444")
        r += 1
    r += 1  # 空一行

    header_row = r
    for c, name in enumerate(REVIEW_HEADER, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # 数据行；加权总分列(R=18)写真实公式，引用 M..Q(13..17)
    for i, row in enumerate(rows):
        excel_row = header_row + 1 + i
        for c, val in enumerate(row, start=1):
            ws.cell(row=excel_row, column=c, value=val)
        ws.cell(row=excel_row, column=18,
                value=(f"=0.30*M{excel_row}+0.25*N{excel_row}+0.20*O{excel_row}"
                       f"+0.15*P{excel_row}+0.10*Q{excel_row}"))

    widths = {"A": 8, "B": 34, "C": 16, "D": 12, "E": 14, "F": 18,
              "G": 20, "H": 20, "I": 18, "J": 16, "K": 16, "L": 16,
              "M": 11, "N": 11, "O": 12, "P": 11, "Q": 11, "R": 14, "S": 24}
    for col, wdt in widths.items():
        ws.column_dimensions[col].width = wdt
    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(path)
    return True


def write_personalization(path_csv, per_query_top):
    """成对双画像对比子表：同一 text 的 A/B 画像左右并排。"""
    # 按 query_id 去掉尾部 a/b 分组
    groups: dict[str, list[str]] = {}
    for qid, data in per_query_top.items():
        if data["query"].get("special") != "个性化对比":
            continue
        base = qid[:-1] if qid[-1] in ("a", "b") else qid
        groups.setdefault(base, []).append(qid)

    def outfit_str(extracted):
        if not extracted:
            return "—"
        cols = extracted["display_columns"]
        return " | ".join(f"{c}:{cols[c]}" for c in E.SLOT_COLUMNS if cols[c] != "—")

    header = ["对比组", "query文本", "场景", "季节/温度",
              "画像A", "A的搭配", "画像B", "B的搭配",
              "差异化是否明显(1-5)", "评审备注"]
    guide = ["【个性化对比说明】同一 query 文本、不同用户画像，左右并排。",
             "只对『两套搭配的差异化是否明显、且各自贴合对应画像』打分：5=差异显著且各自贴合；3=有一定差异；1=几乎相同/千人一面。"]

    with open(path_csv, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        for line in guide:
            w.writerow([line])
        w.writerow([])
        w.writerow(header)
        for base in sorted(groups):
            ids = sorted(groups[base])
            if len(ids) < 2:
                continue
            a, b = ids[0], ids[1]
            qa, qb = per_query_top[a], per_query_top[b]
            w.writerow([
                base, qa["query"]["text"], qa["query"].get("scenario", ""),
                f"{qa['query'].get('season','')} / {qa['query'].get('temp_range','')}",
                qa["query"].get("profile_variant", ""), outfit_str(qa["top"]),
                qb["query"].get("profile_variant", ""), outfit_str(qb["top"]),
                "", "",
            ])


def main():
    ap = argparse.ArgumentParser(description="Stylee 离线搭配能力评测")
    ap.add_argument("--catalog", default="catalog.sample.json")
    ap.add_argument("--queries", default="queries.json")
    ap.add_argument("--provider", default="mock", help="mock|deepseek|qwen")
    ap.add_argument("--out", default="results/")
    ap.add_argument("--n", type=int, default=4, help="每条 query 请求几套")
    args = ap.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    catalog_path = args.catalog if os.path.isabs(args.catalog) else os.path.join(here, args.catalog)
    queries_path = args.queries if os.path.isabs(args.queries) else os.path.join(here, args.queries)
    out_dir = args.out if os.path.isabs(args.out) else os.path.join(here, args.out)
    os.makedirs(out_dir, exist_ok=True)

    catalog = E.load_json(catalog_path)
    queries = E.load_json(queries_path)
    print(f"[eval] catalog={len(catalog)} 件, queries={len(queries)} 条, provider={args.provider}")

    try:
        provider = E.make_provider(args.provider)
    except E.ProviderError as e:
        print(f"[error] provider '{args.provider}' 初始化失败（多半是缺 API key）：{e}")
        print("        离线可复现请用 --provider mock；真实模型需配置对应环境变量后重跑。")
        sys.exit(2)

    jsonl_records, review_rows, per_query_top = build_rows(queries, catalog, provider, args.n)

    jsonl_path = os.path.join(out_dir, "outfits.jsonl")
    csv_path = os.path.join(out_dir, "review_sheet.csv")
    xlsx_path = os.path.join(out_dir, "review_sheet.xlsx")
    pers_path = os.path.join(out_dir, "review_sheet_personalization.csv")

    write_jsonl(jsonl_path, jsonl_records)
    write_review_csv(csv_path, review_rows)
    xlsx_ok = write_review_xlsx(xlsx_path, review_rows)
    write_personalization(pers_path, per_query_top)

    print(f"[eval] 完成 ✅  产物写入 {out_dir}")
    print(f"       - outfits.jsonl                    ({len(jsonl_records)} 行)")
    print(f"       - review_sheet.csv                 ({len(review_rows)} 条待评)")
    print(f"       - review_sheet.xlsx                {'(含公式)' if xlsx_ok else '(跳过)'}")
    print(f"       - review_sheet_personalization.csv (成对双画像)")


if __name__ == "__main__":
    main()
