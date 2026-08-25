#!/usr/bin/env python3
"""Stylee 线上真实模型服务搭配评测（远程 /recommend 版）。

与 run_eval.py 的区别：run_eval.py 是进程内跑 pipeline；本脚本真实 POST 到
线上服务 {STYLEE_API}/recommend（默认 Render 部署），带 Supabase Bearer token。

流程：
  catalog.json + queries.json
    -> eval_lib.build_payload(query, catalog, n) 构造 App JSON
    -> POST {STYLEE_API}/recommend (Authorization: Bearer <supabase access_token>)
    -> 解析 outfits(owned_item_ids/recommended_items/comment/name)
    -> results_real/outfits.jsonl + review_sheet.csv/.xlsx
    -> 打印校验统计
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
import urllib.error
import urllib.request

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import eval_lib as E          # noqa: E402
import _auth                   # noqa: E402

STYLEE_API = os.environ.get("STYLEE_API", "https://stylee-model-service.onrender.com").rstrip("/")

# 硬校验用的品类判定
BASE_TOP = {"上装"}
BASE_BOTTOM = {"下装"}
BASE_DRESS = {"连衣裙", "连体装"}
BASE_SHOES = {"鞋", "鞋履"}

REVIEW_HEADER = [
    "编号", "题目query", "分层", "场景", "风格",
    "模型搭配结果", "硬校验", "平铺拼图",
]

GUIDE_LINES = [
    "【线上真实模型评测表】数据来自 Stylee 线上服务 /recommend（provider=deepseek），非 mock。",
    "模型搭配结果：每道题模型从 748 件真实单品池搭出的整套（owned 单品名+item_id），含补买建议。",
    "硬校验规则：通过=覆盖(上装+下装或连衣裙)且有鞋且所有 item_id 真实存在于单品池；缺件=缺基础层或缺鞋；不合规=存在池中不存在的 item_id。",
    "平铺拼图：单品去背 PNG 拼图占位（图片受下载权限限制时留空）。",
]


def post_recommend(payload: dict, token: str, timeout: int = 150, retries: int = 2):
    body = json.dumps(payload).encode("utf-8")
    last_err = None
    for attempt in range(retries + 1):
        req = urllib.request.Request(
            STYLEE_API + "/recommend",
            data=body,
            headers={"Authorization": "Bearer " + token,
                     "Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return json.loads(resp.read().decode("utf-8")), None
        except urllib.error.HTTPError as e:
            detail = ""
            try:
                detail = e.read().decode("utf-8")[:300]
            except Exception:
                pass
            last_err = f"HTTP {e.code}: {detail}"
        except Exception as e:  # noqa: BLE001
            last_err = f"{type(e).__name__}: {e}"
        if attempt < retries:
            time.sleep(3 + attempt * 3)
    return None, last_err


def hard_check(owned_ids: list[str], cat_idx: dict) -> tuple[str, dict]:
    """对单套搭配做硬校验，返回 (结论, 详情)。"""
    cats = []
    invalid = []
    for iid in owned_ids:
        it = cat_idx.get(iid)
        if it is None:
            invalid.append(iid)
        else:
            cats.append(it.get("category", ""))
    catset = set(cats)
    has_top = bool(catset & BASE_TOP)
    has_bottom = bool(catset & BASE_BOTTOM)
    has_dress = bool(catset & BASE_DRESS)
    has_shoes = bool(catset & BASE_SHOES)
    covers_base = (has_top and has_bottom) or has_dress
    detail = {
        "has_top": has_top, "has_bottom": has_bottom, "has_dress": has_dress,
        "has_shoes": has_shoes, "covers_base": covers_base, "invalid_ids": invalid,
    }
    if not owned_ids:
        return "缺件", detail
    if invalid:
        return "不合规", detail
    if covers_base and has_shoes:
        return "通过", detail
    return "缺件", detail


def parse_outfit(o: dict, cat_idx: dict) -> dict:
    slots = []
    for iid in o.get("owned_item_ids") or []:
        it = cat_idx.get(iid)
        slots.append({
            "owned": True,
            "item_id": iid,
            "name": (it or {}).get("name", "") if it else "",
            "category": (it or {}).get("category", "") if it else "",
            "in_catalog": it is not None,
        })
    gaps = []
    for g in o.get("recommended_items") or []:
        gaps.append({
            "owned": False,
            "name": g.get("name", ""),
            "category": g.get("category", ""),
            "color": g.get("color", ""),
            "reason": g.get("description", ""),
        })
    verdict, detail = hard_check(o.get("owned_item_ids") or [], cat_idx)
    return {
        "name": o.get("name", ""),
        "owned_item_ids": o.get("owned_item_ids") or [],
        "slots": slots,
        "gaps": gaps,
        "reasoning": o.get("comment", ""),
        "hard_check": verdict,
        "hard_check_detail": detail,
    }


def outfit_readable(parsed: dict) -> str:
    parts = []
    for s in parsed["slots"]:
        tag = "" if s["in_catalog"] else "⚠不在池中"
        parts.append(f"{s['category']}:{s['name']}({s['item_id']}){tag}")
    line = " / ".join(parts) if parts else "（无 owned 单品）"
    if parsed["gaps"]:
        gap_txt = "；".join(f"补:{g['name']}" for g in parsed["gaps"])
        line += f"  [建议购买: {gap_txt}]"
    return line


def model_result_text(parsed_outfits: list[dict]) -> str:
    """把全部套拼成可读文本（每套一行）。"""
    lines = []
    for i, p in enumerate(parsed_outfits, 1):
        occ = p["name"] or f"方案{i}"
        lines.append(f"套{i}·{occ}｜{outfit_readable(p)}｜理由:{p['reasoning']}")
    return "\n".join(lines) if lines else "（模型未返回搭配）"


def write_jsonl(path, records):
    with open(path, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def write_review_csv(path, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        for line in GUIDE_LINES:
            w.writerow([line])
        w.writerow([])
        w.writerow(REVIEW_HEADER)
        for r in rows:
            w.writerow(r)


def write_review_xlsx(path, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception as e:  # pragma: no cover
        print(f"[warn] 无 openpyxl，跳过 xlsx：{e}")
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "线上真实评测"
    r = 1
    for line in GUIDE_LINES:
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(size=10, color="444444")
        r += 1
    r += 1
    header_row = r
    for c, name in enumerate(REVIEW_HEADER, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for i, row in enumerate(rows):
        er = header_row + 1 + i
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=er, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {"A": 8, "B": 34, "C": 8, "D": 16, "E": 14, "F": 80, "G": 10, "H": 12}
    for col, wdt in widths.items():
        ws.column_dimensions[col].width = wdt
    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(path)
    return True


def main():
    ap = argparse.ArgumentParser(description="Stylee 线上真实模型搭配评测")
    ap.add_argument("--catalog", default="catalog.json")
    ap.add_argument("--queries", default="queries.json")
    ap.add_argument("--out", default="results_real/")
    ap.add_argument("--n", type=int, default=4)
    ap.add_argument("--limit", type=int, default=0, help="只跑前 N 条（调试用，0=全部）")
    ap.add_argument("--sleep", type=float, default=1.5, help="请求间隔秒")
    ap.add_argument("--timeout", type=int, default=150)
    args = ap.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    catalog = E.load_json(os.path.join(here, args.catalog))
    queries = E.load_json(os.path.join(here, args.queries))
    if args.limit:
        queries = queries[:args.limit]
    out_dir = os.path.join(here, args.out)
    os.makedirs(out_dir, exist_ok=True)
    cat_idx = E.catalog_index(catalog)

    print(f"[eval-remote] API={STYLEE_API} catalog={len(catalog)} queries={len(queries)} n={args.n}")
    token = _auth.get_access_token()
    token_ts = time.time()
    print(f"[eval-remote] 已获取 access_token (len={len(token)})")

    jsonl_records = []
    review_rows = []
    failures = []
    stats = {"empty": 0, "owned_total": 0, "owned_in_catalog": 0,
             "verdict": {"通过": 0, "缺件": 0, "不合规": 0, "失败": 0}}

    for idx, q in enumerate(queries):
        qid = q["query_id"]
        # token 每小时过期，跑到 ~50min 主动刷新
        if time.time() - token_ts > 3000:
            token = _auth.get_access_token()
            token_ts = time.time()
            print("[eval-remote] token 已刷新")
        payload = E.build_payload(q, catalog, n=args.n)
        t0 = time.time()
        data, err = post_recommend(payload, token, timeout=args.timeout)
        elapsed = round(time.time() - t0, 1)

        if err is not None:
            print(f"[{idx+1}/{len(queries)}] {qid} ✗ 失败({elapsed}s): {err}")
            failures.append({"query_id": qid, "error": err})
            stats["verdict"]["失败"] += 1
            review_rows.append([
                qid, q.get("text", ""), q.get("tier", ""), q.get("scenario", ""),
                q.get("style", ""), f"（请求失败：{err}）", "失败", "",
            ])
            jsonl_records.append({"query_id": qid, "text": q.get("text", ""),
                                  "error": err, "outfits": []})
            continue

        outfits = data.get("outfits") or []
        parsed = [parse_outfit(o, cat_idx) for o in outfits]
        if not parsed:
            stats["empty"] += 1
        # 统计 owned 命中
        for p in parsed:
            for s in p["slots"]:
                stats["owned_total"] += 1
                if s["in_catalog"]:
                    stats["owned_in_catalog"] += 1
        # 硬校验以首套为准
        primary_verdict = parsed[0]["hard_check"] if parsed else "缺件"
        stats["verdict"][primary_verdict] += 1

        result_text = model_result_text(parsed)
        review_rows.append([
            qid, q.get("text", ""), q.get("tier", ""), q.get("scenario", ""),
            q.get("style", ""), result_text, primary_verdict, "",
        ])
        jsonl_records.append({
            "query_id": qid,
            "text": q.get("text", ""),
            "labels": {k: q.get(k, "") for k in
                       ("scenario", "style", "color_system", "season",
                        "temp_range", "gender", "special", "profile_variant",
                        "difficulty", "tier")},
            "payload_weather": payload["weather"],
            "payload_profile": payload["profile"],
            "provider": (data.get("trace") or {}).get("provider"),
            "num_outfits": len(parsed),
            "primary_hard_check": primary_verdict,
            "outfits": [
                {
                    "rank": i + 1,
                    "occasion": p["name"],
                    "owned_item_ids": p["owned_item_ids"],
                    "slots": p["slots"],
                    "gaps": p["gaps"],
                    "reasoning": p["reasoning"],
                    "hard_check": p["hard_check"],
                    "hard_check_detail": p["hard_check_detail"],
                }
                for i, p in enumerate(parsed)
            ],
            "trace": data.get("trace"),
        })
        print(f"[{idx+1}/{len(queries)}] {qid} ✓ {elapsed}s 套数={len(parsed)} 硬校验={primary_verdict}")
        time.sleep(args.sleep)

    # 落盘
    jsonl_path = os.path.join(out_dir, "outfits.jsonl")
    csv_path = os.path.join(out_dir, "review_sheet.csv")
    xlsx_path = os.path.join(out_dir, "review_sheet.xlsx")
    write_jsonl(jsonl_path, jsonl_records)
    write_review_csv(csv_path, review_rows)
    xlsx_ok = write_review_xlsx(xlsx_path, review_rows)
    if failures:
        with open(os.path.join(out_dir, "failures.json"), "w", encoding="utf-8") as f:
            json.dump(failures, f, ensure_ascii=False, indent=2)

    hit_ratio = (stats["owned_in_catalog"] / stats["owned_total"] * 100) if stats["owned_total"] else 0
    print("\n==== 校验统计 ====")
    print(f"总题数: {len(queries)}  成功: {len(queries)-len(failures)}  失败: {len(failures)}")
    print(f"空结果题数: {stats['empty']}")
    print(f"owned 选品总数: {stats['owned_total']}  命中 catalog: {stats['owned_in_catalog']}  命中率: {hit_ratio:.2f}%")
    print(f"硬校验(首套)分布: {stats['verdict']}")
    print(f"产物: {jsonl_path}\n      {csv_path}\n      {xlsx_path}{'' if xlsx_ok else '(xlsx跳过)'}")
    if failures:
        print(f"失败题号: {[f['query_id'] for f in failures]}")


if __name__ == "__main__":
    main()
